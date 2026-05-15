"""XLSX / CSV pipeline-state exporter.

This is the customer-facing artefact when a customer's deliverable is
spreadsheet-only. The exporter walks the CRM adapter, pulls every record
of each kind, and produces:

- One ``pipeline_<customer>_<date>.xlsx`` workbook with one sheet per
  record kind (Targets, Stakeholders, Signals, Activities) plus a
  Pending_Approvals sheet.
- One ``pipeline_<customer>_<date>.csv.zip`` archive containing the same
  data as one CSV per record kind, for customers who prefer CSV.
- An ``approvals_inbox.csv`` the customer can edit in place. On the next
  agent wake, the Hand-off agent reads this file and applies the
  decisions in the ``decision`` column ("approve" / "reject" /
  blank = still pending).

The exporter is platform-level — its column layout for each record kind
comes from the vertical pack's ``schemas/*.yaml`` files (the same source
of truth the agents themselves use). Customer-specific styling (brand
colours, logo) comes from ``customers/<customer>/brand.yaml``.
"""

from __future__ import annotations

import csv
import io
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from nba_platform.integrations.crm.base import CrmAdapter, RecordKind


@dataclass
class ExportResult:
    """Where the exported artefacts landed on disk."""

    xlsx_path: Path
    csv_zip_path: Path
    approvals_inbox_path: Path
    record_counts: dict[str, int]


# The four record kinds always export. Order is the order they appear as
# workbook sheets / CSV files.
EXPORT_KINDS: tuple[RecordKind, ...] = (
    RecordKind.TARGET_ORGANISATION,
    RecordKind.STAKEHOLDER,
    RecordKind.SIGNAL,
    RecordKind.ACTIVITY,
)


class PipelineExporter:
    """Exports the current pipeline state to XLSX + CSV.

    ``vertical_pack`` is the dict produced by the config loader — its
    ``schemas`` section provides the field ordering for each sheet.
    ``brand`` is the customer's brand.yaml as a dict — used for header
    styling.
    """

    def __init__(
        self,
        *,
        customer: str,
        output_dir: Path,
        vertical_pack: dict[str, Any],
        brand: dict[str, Any],
    ) -> None:
        self._customer = customer
        self._output_dir = Path(output_dir)
        self._output_dir.mkdir(parents=True, exist_ok=True)
        self._vertical_pack = vertical_pack
        self._brand = brand
        self._timestamp = datetime.now(timezone.utc)

    async def export(self, crm: CrmAdapter) -> ExportResult:
        """Pull state from the CRM adapter, render workbook + CSVs + approvals inbox."""
        rows_by_kind: dict[RecordKind, list[dict[str, Any]]] = {}
        record_counts: dict[str, int] = {}

        for kind in EXPORT_KINDS:
            hits = await crm.search(kind, query={}, limit=100_000)
            rows = [h.fields for h in hits]
            rows_by_kind[kind] = rows
            record_counts[kind.value] = len(rows)

        pending = await crm.list_pending_approvals(limit=10_000)
        pending_rows = [
            {
                "approval_id": p.approval_id,
                "record_kind": p.record_kind.value,
                "natural_key": (p.proposed_fields or {}).get("natural_key", ""),
                "proposed_fields_summary": _summarise(p.proposed_fields),
                "rationale": p.rationale,
                "proposed_by_agent": p.proposed_by_agent,
                "created_at": p.created_at.isoformat(),
                "decision": "",  # customer fills in: approve / reject
                "decision_note": "",
            }
            for p in pending
        ]
        record_counts["pending_approvals"] = len(pending_rows)

        date_tag = self._timestamp.strftime("%Y%m%d")
        xlsx_path = self._output_dir / f"pipeline_{self._customer}_{date_tag}.xlsx"
        csv_zip_path = self._output_dir / f"pipeline_{self._customer}_{date_tag}.csv.zip"
        approvals_inbox_path = self._output_dir / "approvals_inbox.csv"

        self._write_xlsx(xlsx_path, rows_by_kind, pending_rows)
        self._write_csv_zip(csv_zip_path, rows_by_kind, pending_rows)
        self._write_approvals_inbox(approvals_inbox_path, pending_rows)

        return ExportResult(
            xlsx_path=xlsx_path,
            csv_zip_path=csv_zip_path,
            approvals_inbox_path=approvals_inbox_path,
            record_counts=record_counts,
        )

    # ------------------------------------------------------------------
    # XLSX writing
    # ------------------------------------------------------------------

    def _write_xlsx(
        self,
        path: Path,
        rows_by_kind: dict[RecordKind, list[dict[str, Any]]],
        pending_rows: list[dict[str, Any]],
    ) -> None:
        wb = Workbook()
        # Remove the default sheet — we'll add named sheets.
        default = wb.active
        wb.remove(default)

        for kind in EXPORT_KINDS:
            sheet_name = _sheet_name_for_kind(kind)
            columns = self._columns_for_kind(kind)
            rows = rows_by_kind[kind]
            ws = wb.create_sheet(sheet_name)
            self._write_sheet(ws, columns, rows)

        ws_pending = wb.create_sheet("Pending_Approvals")
        pending_columns = [
            "approval_id",
            "record_kind",
            "natural_key",
            "proposed_fields_summary",
            "rationale",
            "proposed_by_agent",
            "created_at",
            "decision",
            "decision_note",
        ]
        self._write_sheet(ws_pending, pending_columns, pending_rows, highlight_decision=True)

        wb.save(path)

    def _write_sheet(
        self,
        ws: Any,
        columns: list[str],
        rows: list[dict[str, Any]],
        *,
        highlight_decision: bool = False,
    ) -> None:
        header_fill = PatternFill(
            "solid", fgColor=_hex(self._brand.get("colours", {}).get("primary", "#0B1F3A"))
        )
        header_font = Font(bold=True, color="FFFFFF")
        ws.append(columns)
        for col_idx, _ in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")

        for row in rows:
            ws.append([_render_cell(row.get(col, "")) for col in columns])

        # Auto-ish column widths
        for col_idx, col_name in enumerate(columns, start=1):
            letter = get_column_letter(col_idx)
            max_len = max(
                [len(col_name)]
                + [len(_render_cell(row.get(col_name, ""))) for row in rows[:200]]
            )
            ws.column_dimensions[letter].width = min(max_len + 2, 60)

        ws.freeze_panes = "A2"

        if highlight_decision:
            # Make the decision column visually obvious so the customer
            # sees where to type.
            decision_col_idx = columns.index("decision") + 1
            highlight = PatternFill("solid", fgColor="FFF2CC")
            for row_idx in range(2, len(rows) + 2):
                ws.cell(row=row_idx, column=decision_col_idx).fill = highlight

    def _columns_for_kind(self, kind: RecordKind) -> list[str]:
        """Column ordering driven by the vertical pack's schema for this kind.

        Falls back to a generic ordering when the schema doesn't pin one.
        """
        schemas = self._vertical_pack.get("schemas", {})
        schema = schemas.get(kind.value, {})
        fields = schema.get("fields", {})
        if fields and isinstance(fields, dict):
            return list(fields.keys())
        # Fallback — every kind shares these minimal columns.
        return ["natural_key", "created_at", "updated_at"]

    # ------------------------------------------------------------------
    # CSV / approvals-inbox writing
    # ------------------------------------------------------------------

    def _write_csv_zip(
        self,
        path: Path,
        rows_by_kind: dict[RecordKind, list[dict[str, Any]]],
        pending_rows: list[dict[str, Any]],
    ) -> None:
        with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
            for kind in EXPORT_KINDS:
                columns = self._columns_for_kind(kind)
                rows = rows_by_kind[kind]
                zf.writestr(
                    f"{_sheet_name_for_kind(kind).lower()}.csv",
                    _rows_to_csv(columns, rows),
                )
            pending_columns = [
                "approval_id",
                "record_kind",
                "natural_key",
                "proposed_fields_summary",
                "rationale",
                "proposed_by_agent",
                "created_at",
                "decision",
                "decision_note",
            ]
            zf.writestr(
                "pending_approvals.csv",
                _rows_to_csv(pending_columns, pending_rows),
            )

    def _write_approvals_inbox(
        self,
        path: Path,
        pending_rows: list[dict[str, Any]],
    ) -> None:
        # The inbox file persists between runs. We write fresh each time
        # — the agent's approval-resolver reads the previous version
        # before this one is rewritten (the read/write sequence is the
        # Hand-off agent's responsibility, not the exporter's).
        columns = [
            "approval_id",
            "record_kind",
            "natural_key",
            "proposed_fields_summary",
            "rationale",
            "proposed_by_agent",
            "created_at",
            "decision",
            "decision_note",
        ]
        with path.open("w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            for row in pending_rows:
                writer.writerow({c: row.get(c, "") for c in columns})


# ----------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------


def _sheet_name_for_kind(kind: RecordKind) -> str:
    return {
        RecordKind.TARGET_ORGANISATION: "Targets",
        RecordKind.STAKEHOLDER: "Stakeholders",
        RecordKind.SIGNAL: "Signals",
        RecordKind.ACTIVITY: "Activities",
    }[kind]


def _hex(s: str) -> str:
    return s.lstrip("#")


def _render_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return ", ".join(_render_cell(v) for v in value)
    if isinstance(value, dict):
        return "; ".join(f"{k}={_render_cell(v)}" for k, v in value.items())
    return str(value)


def _summarise(fields: dict[str, Any]) -> str:
    """One-line summary for the Pending_Approvals row."""
    if not fields:
        return ""
    keys_in_order = (
        "legal_name",
        "full_name",
        "headline",
        "domain",
        "work_email",
        "role_title",
    )
    parts: list[str] = []
    for k in keys_in_order:
        if k in fields and fields[k]:
            parts.append(f"{k}={fields[k]}")
        if len(parts) >= 3:
            break
    if not parts:
        # Fallback — first three fields, whatever they are.
        parts = [f"{k}={v}" for k, v in list(fields.items())[:3]]
    return "; ".join(parts)


def _rows_to_csv(columns: list[str], rows: list[dict[str, Any]]) -> str:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=columns)
    writer.writeheader()
    for row in rows:
        writer.writerow({c: _render_cell(row.get(c, "")) for c in columns})
    return buf.getvalue()

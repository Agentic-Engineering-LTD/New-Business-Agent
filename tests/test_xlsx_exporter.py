"""Tests for the XLSX/CSV pipeline exporter.

Runs the full export against an in-memory-style SQLite adapter (file in
``tmp_path``), then verifies:

- The XLSX workbook is produced with the expected sheets.
- The CSV zip contains a CSV per record kind plus pending_approvals.csv.
- The approvals_inbox.csv exists with the expected columns and any
  pending-approval rows.
"""

from __future__ import annotations

import csv
import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from nba_platform.integrations.crm.base import RecordKind
from nba_platform.integrations.crm.sqlite import SqliteAdapter
from nba_platform.outputs.xlsx_exporter import PipelineExporter


def _build_adapter(tmp_path: Path) -> SqliteAdapter:
    return SqliteAdapter(
        {
            "kind": "sqlite",
            "db_path": str(tmp_path / "pipeline.db"),
            "tables": {
                "target_organisation": "targets",
                "stakeholder": "stakeholders",
                "signal": "signals",
                "activity": "activities",
            },
            "natural_keys": {
                "target_organisation": "domain",
                "stakeholder": "work_email",
            },
            "require_approval": {
                "target_organisation": False,
                "stakeholder": True,
                "signal": False,
                "activity": False,
            },
        }
    )


def _vertical_pack_stub() -> dict:
    """Minimal vertical pack shape — just enough for column ordering."""
    return {
        "schemas": {
            "target_organisation": {
                "fields": {
                    "legal_name": {"type": "string"},
                    "domain": {"type": "string"},
                    "hq_country": {"type": "string"},
                }
            },
            "stakeholder": {
                "fields": {
                    "full_name": {"type": "string"},
                    "work_email": {"type": "string"},
                    "role_band": {"type": "string"},
                }
            },
            "signal": {"fields": {"headline": {"type": "string"}}},
            "activity": {"fields": {"kind": {"type": "string"}}},
        },
        "data_sources": {},
        "ranking": {},
    }


def _brand_stub() -> dict:
    return {"colours": {"primary": "#0B1F3A"}}


@pytest.fixture
async def populated_adapter(tmp_path: Path) -> SqliteAdapter:
    adapter = _build_adapter(tmp_path)
    # Targets
    await adapter.upsert(
        RecordKind.TARGET_ORGANISATION,
        {"legal_name": "Acme Retail", "domain": "acme.example", "hq_country": "GB"},
        natural_key="acme.example",
    )
    await adapter.upsert(
        RecordKind.TARGET_ORGANISATION,
        {"legal_name": "Globex Group", "domain": "globex.example", "hq_country": "DE"},
        natural_key="globex.example",
    )
    # Stakeholder upsert is approval-gated — that produces a pending row.
    await adapter.upsert(
        RecordKind.STAKEHOLDER,
        {
            "full_name": "Jane Doe",
            "work_email": "jane@acme.example",
            "role_band": "head_of_category",
        },
        natural_key="jane@acme.example",
    )
    return adapter


async def test_export_produces_all_three_artefacts(
    populated_adapter: SqliteAdapter, tmp_path: Path
) -> None:
    exporter = PipelineExporter(
        customer="powerplay",
        output_dir=tmp_path / "outputs",
        vertical_pack=_vertical_pack_stub(),
        brand=_brand_stub(),
    )
    result = await exporter.export(populated_adapter)

    assert result.xlsx_path.exists()
    assert result.csv_zip_path.exists()
    assert result.approvals_inbox_path.exists()
    assert result.record_counts["target_organisation"] == 2
    assert result.record_counts["stakeholder"] == 0, (
        "stakeholder was approval-gated, so the table should be empty"
    )
    assert result.record_counts["pending_approvals"] == 1


async def test_xlsx_has_expected_sheets(
    populated_adapter: SqliteAdapter, tmp_path: Path
) -> None:
    exporter = PipelineExporter(
        customer="powerplay",
        output_dir=tmp_path / "outputs",
        vertical_pack=_vertical_pack_stub(),
        brand=_brand_stub(),
    )
    result = await exporter.export(populated_adapter)
    wb = load_workbook(result.xlsx_path)
    expected = {"Targets", "Stakeholders", "Signals", "Activities", "Pending_Approvals"}
    assert expected.issubset(set(wb.sheetnames))

    targets_sheet = wb["Targets"]
    # Header row + 2 data rows
    rows = list(targets_sheet.iter_rows(values_only=True))
    assert rows[0] == ("legal_name", "domain", "hq_country")
    data_domains = sorted(row[1] for row in rows[1:])
    assert data_domains == ["acme.example", "globex.example"]


async def test_csv_zip_contains_one_csv_per_kind(
    populated_adapter: SqliteAdapter, tmp_path: Path
) -> None:
    exporter = PipelineExporter(
        customer="powerplay",
        output_dir=tmp_path / "outputs",
        vertical_pack=_vertical_pack_stub(),
        brand=_brand_stub(),
    )
    result = await exporter.export(populated_adapter)
    with zipfile.ZipFile(result.csv_zip_path) as zf:
        names = set(zf.namelist())
    expected = {"targets.csv", "stakeholders.csv", "signals.csv", "activities.csv", "pending_approvals.csv"}
    assert expected.issubset(names)


async def test_approvals_inbox_has_decision_column(
    populated_adapter: SqliteAdapter, tmp_path: Path
) -> None:
    exporter = PipelineExporter(
        customer="powerplay",
        output_dir=tmp_path / "outputs",
        vertical_pack=_vertical_pack_stub(),
        brand=_brand_stub(),
    )
    result = await exporter.export(populated_adapter)
    with result.approvals_inbox_path.open("r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    assert "decision" in (reader.fieldnames or []), (
        "approvals_inbox.csv must expose a 'decision' column for customer edits"
    )
    assert len(rows) == 1, "should be one pending stakeholder approval"
    assert rows[0]["decision"] == "", "decision column starts blank"
